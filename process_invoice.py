#!/usr/bin/env python3
"""
Invoice PDF → Excel 변환기
- PDF 드래그 앤 드롭 (또는 인자로 경로 전달)
- 배송지 주소에서 리테일러 자동 감지 (Lotte/Shilla/Shinsegae/Hyundai/JDC)
- EAN 기반 SKU 마스터 조회 → 해당 리테일러 코드 매핑
- Output: 엑셀파일 (EAN, Retailer Code, Description, Qty, Unit Price, Total Price)

Usage:
  python3 process_invoice.py path/to/invoice.pdf
  python3 process_invoice.py  # 기본: 파일 선택 다이얼로그 (tkinter)
"""

import os
import sys
import re
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─── Config ────────────────────────────────────────────────
CONFIG_DIR = os.path.expanduser("~/hermes work/po-process")
CONFIG_FILE = os.path.join(CONFIG_DIR, "config.json")

RETAILER_MAP = {
    'lotte':   {'keywords': ['LOTTE'],                  'label': 'Lotte Code',        'file_key': 'lotte'},
    'shilla':  {'keywords': ['SHILLA', '신라'],          'label': 'Shilla Code',       'file_key': 'shilla'},
    'ssg':     {'keywords': ['SHINSEGAE', 'SSG'],       'label': 'SSG Code',          'file_key': 'ssg'},
    'hyundai': {'keywords': ['HYUNDAI', '현대'],         'label': 'Hyundai Code',      'file_key': 'hyundai'},
    'jdc':     {'keywords': ['JEJU FREE', 'JDC'],       'label': 'JDC Code',          'file_key': 'jdc'},
}

# ───────────────────────────────────────────────────────────
import json

def resolve_sku_master(cli_path=None):
    """SKU 마스터 파일 경로 결정: CLI 인자 > 설정파일 > Downloads 최신순"""
    if cli_path and os.path.exists(cli_path):
        return cli_path

    # 설정 파일 확인
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                cfg = json.load(f)
            if 'sku_master_path' in cfg and os.path.exists(cfg['sku_master_path']):
                return cfg['sku_master_path']
        except:
            pass

    # Downloads에서 가장 최근 SKU master file 찾기
    downloads = os.path.expanduser("~/Downloads")
    candidates = []
    if os.path.isdir(downloads):
        for f in os.listdir(downloads):
            if f.startswith('SKU master file') and f.endswith('.xlsx'):
                full = os.path.join(downloads, f)
                candidates.append((os.path.getmtime(full), full))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][1]

    return None


def resolve_output_dir(cli_path=None):
    """출력 디렉토리 결정: CLI 인자 > 설정파일 > ~/Desktop"""
    if cli_path and os.path.isdir(cli_path):
        return cli_path
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                cfg = json.load(f)
            if 'output_dir' in cfg and os.path.isdir(cfg['output_dir']):
                return cfg['output_dir']
        except:
            pass
    return os.path.expanduser("~/Desktop")


def set_output_dir(path):
    """출력 디렉토리를 config에 저장"""
    path = os.path.abspath(path)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    cfg = {}
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r') as f:
                cfg = json.load(f)
        except:
            pass
    cfg['output_dir'] = path
    with open(CONFIG_FILE, 'w') as f:
        json.dump(cfg, f, indent=2)
    return path

# ───────────────────────────────────────────────────────────

def detect_retailer(pdf_path):
    """PDF 배송지 주소를 읽어 리테일러 감지"""
    import fitz
    doc = fitz.open(pdf_path)
    page = doc[0]
    words = page.get_text("words")

    # 오른쪽 영역(x>400)에서 "DELIVERY ADDRESS" 아래 텍스트 수집 (y: 20~180)
    delivery_text = []
    for w in words:
        x0, y0, x1, y1, word = w[:5]
        if x0 > 400 and 18 <= y0 <= 180:
            delivery_text.append((y0, x0, word))

    # y 기준 정렬 후 텍스트 조합
    delivery_text.sort(key=lambda t: (t[0], t[1]))
    combined = ' '.join([w for _, _, w in delivery_text]).upper()

    doc.close()

    # 키워드 매칭 (순서: Lotte > Shilla > Shinsegae > Hyundai > JDC)
    for retailer_id, info in RETAILER_MAP.items():
        for kw in info['keywords']:
            if kw in combined:
                print(f"  🏪 리테일러 감지: {info['label']} ({kw})")
                return retailer_id

    print("  ⚠️  리테일러를 감지할 수 없습니다. 기본값: lotte")
    return 'lotte'


def load_sku_master(path):
    """SKU master file 로드 → {ean_str: {lotte, shilla, ssg, hyundai, jdc, desc}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if name not in ('Sheet1',):
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 헤더 찾기: "EAN Code" 행
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        for cell in row:
            if cell.value and 'EAN' in str(cell.value) and 'Code' in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        print("❌ SKU master에서 헤더 행을 찾을 수 없습니다.")
        sys.exit(1)

    # 헤더 컬럼 매핑
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            val = str(cell.value).strip()
            cl = cell.column_letter
            if 'EAN' in val:
                headers['ean'] = cl
            elif 'Lotte' in val or 'LOTTE' in val:
                headers['lotte'] = cl
            elif 'Shilla' in val or 'SHILLA' in val:
                headers['shilla'] = cl
            elif 'SSG' in val:
                headers['ssg'] = cl
            elif 'Hyundai' in val or 'HYUNDAI' in val:
                headers['hyundai'] = cl
            elif 'JDC' in val:
                headers['jdc'] = cl
            elif 'Product Label' in val or 'Label' in val:
                headers['desc'] = cl

    required = ['ean', 'lotte', 'shilla', 'ssg', 'hyundai', 'jdc']
    missing = [r for r in required if r not in headers]
    if missing:
        print(f"❌ SKU master에서 컬럼을 찾을 수 없음: {missing}")
        print(f"   찾은 헤더: {headers}")
        sys.exit(1)

    # 데이터 읽기
    ean_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ean_idx = ord(headers['ean']) - ord('A')
        ean_val = row[ean_idx] if ean_idx < len(row) else None
        if ean_val is None:
            continue
        ean_str = str(int(ean_val)) if isinstance(ean_val, (int, float)) else str(ean_val).strip()

        def get_val(col_letter):
            idx = ord(col_letter) - ord('A')
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else ''

        ean_map[ean_str] = {
            'lotte': get_val(headers['lotte']),
            'shilla': get_val(headers['shilla']),
            'ssg': get_val(headers['ssg']),
            'hyundai': get_val(headers['hyundai']),
            'jdc': get_val(headers['jdc']),
            'desc': get_val(headers.get('desc', 'A')),
        }

    print(f"  📋 SKU master 로드 완료: {len(ean_map)}개 제품")
    return ean_map


def parse_invoice_pdf(pdf_path):
    """PDF에서 인보이스 데이터 추출 (EAN, Description, Qty, UnitPrice, TotalPrice)"""
    import fitz

    doc = fitz.open(pdf_path)
    all_items = []

    for page_num in range(len(doc)):
        page = doc[page_num]
        words = page.get_text("words")

        rows = {}
        for w in words:
            x0, y0, x1, y1, word, block_no, line_no, word_no = w
            row_key = round(y0 / 5) * 5
            if row_key not in rows:
                rows[row_key] = []
            rows[row_key].append((x0, word))

        for y in sorted(rows.keys()):
            items = sorted(rows[y], key=lambda t: t[0])
            words_only = [item[1] for item in items]

            first = words_only[0].replace(',', '')
            if re.match(r'^\d{13}$', first):
                if len(words_only) < 9:
                    continue
                suffix = words_only[-8:]
                desc_words = words_only[1:-8]
                description = ' '.join(desc_words)

                qty_str = suffix[2].replace(',', '')
                up_str = suffix[3].replace(',', '')
                tp_str = suffix[4].replace(',', '')

                try:
                    qty = int(float(qty_str)) if '.' not in qty_str else float(qty_str)
                except:
                    qty = 0
                try:
                    unit_price = float(up_str)
                except:
                    unit_price = 0.0
                try:
                    total_price = float(tp_str)
                except:
                    total_price = 0.0

                all_items.append({
                    'ean': first,
                    'description': description,
                    'quantity': qty,
                    'unit_price': unit_price,
                    'total_price': total_price,
                })

    doc.close()
    print(f"  📄 PDF 파싱 완료: {len(all_items)}개 품목")
    return all_items


def lookup_retailer_codes(items, ean_map, retailer_id):
    """SKU master 조회 → 감지된 리테일러 코드만 추출"""
    matched = 0
    for item in items:
        ean = item['ean']
        if ean in ean_map:
            sku = ean_map[ean]
            item['retailer_code'] = sku[retailer_id]
            item['retailer_desc_sku'] = sku['desc']
            matched += 1
        else:
            item['retailer_code'] = ''
            item['retailer_desc_sku'] = ''

    print(f"  🔗 {RETAILER_MAP[retailer_id]['label']} 매칭: {matched}/{len(items)}")
    return items


def create_excel(items, pdf_path, retailer_id, output_dir=None):
    """결과 엑셀 파일 생성"""
    out_dir = output_dir or resolve_output_dir()
    pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
    output_name = f"{pdf_basename}_processed.xlsx"
    output_path = os.path.join(out_dir, output_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    retailer_label = RETAILER_MAP[retailer_id]['label']

    # ─── 스타일 ────────────────────────────────────────────
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    sum_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sum_font = Font(name='Calibri', bold=True, size=11)

    # ─── 헤더 ─────────────────────────────────────────────
    # 컬럼: 1.EAN | 2.Retailer Code | 3.Description | 4.Quantity | 5.Unit Price | 6.Total Price
    col_defs = [
        ('EAN Number', 22),
        (retailer_label, 18),
        ('Description', 52),
        ('Quantity', 12),
        ('Unit Price (USD)', 16),
        ('Total Price (USD)', 18),
    ]

    for col_idx, (header_text, width) in enumerate(col_defs, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ─── 데이터 ───────────────────────────────────────────
    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item['ean']).font = data_font
        ws.cell(row=row_idx, column=1).alignment = center_alignment

        ws.cell(row=row_idx, column=2, value=item['retailer_code']).font = data_font
        ws.cell(row=row_idx, column=2).alignment = center_alignment

        ws.cell(row=row_idx, column=3, value=item['description']).font = data_font
        ws.cell(row=row_idx, column=3).alignment = data_alignment

        c_qty = ws.cell(row=row_idx, column=4, value=item['quantity'])
        c_qty.font = data_font
        c_qty.alignment = number_alignment
        c_qty.number_format = '#,##0'

        c_up = ws.cell(row=row_idx, column=5, value=item['unit_price'])
        c_up.font = data_font
        c_up.alignment = number_alignment
        c_up.number_format = '#,##0.00'

        c_tp = ws.cell(row=row_idx, column=6, value=item['total_price'])
        c_tp.font = data_font
        c_tp.alignment = number_alignment
        c_tp.number_format = '#,##0.00'

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = thin_border

    # ─── 합계 행 ──────────────────────────────────────────
    sum_row = len(items) + 2
    ws.cell(row=sum_row, column=1, value='TOTAL').font = sum_font
    ws.cell(row=sum_row, column=1).fill = sum_fill
    ws.cell(row=sum_row, column=1).alignment = center_alignment
    ws.cell(row=sum_row, column=1).border = thin_border

    for col in range(2, 4):
        ws.cell(row=sum_row, column=col).fill = sum_fill
        ws.cell(row=sum_row, column=col).border = thin_border

    data_start = 2
    data_end = len(items) + 1

    sum_cols = {
        4: ('#,##0', 'Quantity'),
        5: ('#,##0.00', 'Unit Price'),
        6: ('#,##0.00', 'Total Price'),
    }
    for col, (fmt, _) in sum_cols.items():
        c = ws.cell(row=sum_row, column=col)
        c.value = f'=SUM({get_column_letter(col)}{data_start}:{get_column_letter(col)}{data_end})'
        c.font = sum_font
        c.fill = sum_fill
        c.alignment = number_alignment
        c.border = thin_border
        c.number_format = fmt

    # 서식 마무리
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{data_end}'

    wb.save(output_path)
    print(f"  💾 저장 완료: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Invoice PDF → Excel 변환')
    parser.add_argument('pdf', nargs='?', help='PDF 파일 경로')
    parser.add_argument('--sku-master', help='SKU master 파일 경로 (지정 안하면 config/Downloads 자동 탐색)')
    parser.add_argument('--output-dir', help='저장 폴더 경로 (기본: ~/Desktop 또는 config 설정)')
    args = parser.parse_args()

    pdf_path = args.pdf
    if not pdf_path:
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            pdf_path = filedialog.askopenfilename(
                title='인보이스 PDF 선택',
                filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')]
            )
            root.destroy()
            if not pdf_path:
                print("❌ 파일이 선택되지 않았습니다.")
                sys.exit(1)
        except ImportError:
            print("❌ PDF 파일 경로를 인자로 전달해주세요.")
            print(f"   사용법: python3 {sys.argv[0]} path/to/invoice.pdf")
            sys.exit(1)

    if not os.path.exists(pdf_path):
        print(f"❌ 파일을 찾을 수 없음: {pdf_path}")
        sys.exit(1)
    if not pdf_path.lower().endswith('.pdf'):
        print(f"❌ PDF 파일이 아닙니다: {pdf_path}")
        sys.exit(1)

    # SKU 마스터 경로 결정
    sku_master_path = resolve_sku_master(args.sku_master)
    if not sku_master_path or not os.path.exists(sku_master_path):
        print("❌ SKU master file을 찾을 수 없습니다.")
        print(f"   설정: {CONFIG_FILE}")
        print("   SKU master 파일을 앱에 드래그하거나 Downloads 폴더에 넣어주세요.")
        sys.exit(1)

    # 출력 경로 결정
    output_dir = resolve_output_dir(args.output_dir)

    print(f"\n{'='*50}")
    print(f"  Invoice PDF → Excel 변환")
    print(f"{'='*50}")
    print(f"  📁 PDF: {pdf_path}")
    print(f"  📋 SKU: {os.path.basename(sku_master_path)}")
    print(f"  💾 저장: {output_dir}")

    # 1. 배송지 주소 → 리테일러 감지
    retailer_id = detect_retailer(pdf_path)

    # 2. SKU 마스터 로드
    ean_map = load_sku_master(sku_master_path)

    # 3. PDF 파싱
    items = parse_invoice_pdf(pdf_path)
    if not items:
        print("❌ PDF에서 데이터를 추출할 수 없습니다.")
        sys.exit(1)

    # 4. Retailer 코드 조회
    items = lookup_retailer_codes(items, ean_map, retailer_id)

    # 5. 엑셀 생성
    output_path = create_excel(items, pdf_path, retailer_id, output_dir)

    # 요약
    total_qty = sum(item['quantity'] for item in items)
    total_amount = sum(item['total_price'] for item in items)
    matched = sum(1 for item in items if item['retailer_code'])

    print(f"\n{'='*50}")
    print(f"  ✅ 변환 완료")
    print(f"  🏪 리테일러: {RETAILER_MAP[retailer_id]['label']}")
    print(f"  📊 {len(items)}개 품목 처리")
    print(f"  🔗 {matched}/{len(items)}개 코드 매칭")
    print(f"  📦 총 수량: {total_qty:,}")
    print(f"  💵 총 금액: ${total_amount:,.2f}")
    print(f"  📎 결과: {output_path}")
    print(f"{'='*50}\n")


if __name__ == '__main__':
    main()
