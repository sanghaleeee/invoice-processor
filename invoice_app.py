# Invoice Processor - Windows GUI
# Cross-platform tkinter app. macOS에서 개발/테스트, Windows에서 구동 가능.
#
# 빌드 (Windows):
#   pip install pyinstaller
#   pyinstaller --onefile --windowed --icon=appicon.ico invoice_app.py
#
# 의존성:
#   pip install PyMuPDF openpyxl

import os
import sys
import re
import json
import threading
import subprocess
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ─── Config ────────────────────────────────────────────────
HOME = Path.home()
if sys.platform == 'win32':
    DOWNLOADS = HOME / 'Downloads'
    DESKTOP = HOME / 'Desktop'
    CONFIG_DIR = HOME / 'hermes work' / 'po-process'
else:
    DOWNLOADS = HOME / 'Downloads'
    DESKTOP = HOME / 'Desktop'
    CONFIG_DIR = HOME / 'hermes work' / 'po-process'

CONFIG_FILE = CONFIG_DIR / 'config.json'
SKU_STORAGE = CONFIG_DIR / 'sku-master'

RETAILER_MAP = {
    'lotte':   ('LOTTE', None,                 'Lotte Code'),
    'shilla':  ('SHILLA', '신라',               'Shilla Code'),
    'ssg':     ('SHINSEGAE', 'SSG',            'SSG Code'),
    'hyundai': ('HYUNDAI', '현대',              'Hyundai Code'),
    'jdc':     ('JEJU FREE', 'JDC',            'JDC Code'),
}

# ─── Core Logic (same as process_invoice.py) ──────────────

def resolve_sku_master(cli_path=None):
    if cli_path and os.path.exists(cli_path):
        return cli_path
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r') as f:
                cfg = json.load(f)
            if 'sku_master_path' in cfg and os.path.exists(cfg['sku_master_path']):
                return cfg['sku_master_path']
        except:
            pass
    if DOWNLOADS.is_dir():
        candidates = []
        for f in os.listdir(str(DOWNLOADS)):
            if f.startswith('SKU master file') and f.endswith('.xlsx'):
                full = str(DOWNLOADS / f)
                candidates.append((os.path.getmtime(full), full))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][1]
    return None


def resolve_output_dir(cli_path=None):
    if cli_path and os.path.isdir(cli_path):
        return cli_path
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE) as f:
                cfg = json.load(f)
            if 'output_dir' in cfg and os.path.isdir(cfg['output_dir']):
                return cfg['output_dir']
        except:
            pass
    return str(Path.home() / 'Desktop')


def set_output_dir(path):
    path = os.path.abspath(path)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    cfg = {}
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE) as f:
                cfg = json.load(f)
        except:
            pass
    cfg['output_dir'] = path
    with open(CONFIG_FILE, 'w') as f:
        json.dump(cfg, f, indent=2)
    return path


def detect_retailer(pdf_path):
    if fitz is None:
        return 'lotte'
    doc = fitz.open(pdf_path)
    page = doc[0]
    words = page.get_text("words")
    delivery_text = []
    for w in words:
        x0, y0, x1, y1, word = w[:5]
        if x0 > 400 and 18 <= y0 <= 180:
            delivery_text.append((y0, x0, word))
    delivery_text.sort(key=lambda t: (t[0], t[1]))
    combined = ' '.join([w for _, _, w in delivery_text]).upper()
    doc.close()
    for rid, info in RETAILER_MAP.items():
        for kw in info[:2]:
            if kw and kw in combined:
                return rid
    return 'lotte'


def load_sku_master(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if name not in ('Sheet1',):
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        for cell in row:
            if cell.value and 'EAN' in str(cell.value) and 'Code' in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        raise ValueError("Cannot find header row in SKU master")

    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            val = str(cell.value).strip()
            cl = cell.column_letter
            if 'EAN' in val:
                headers['ean'] = cl
            elif 'Lotte' in val or 'LOTTE' in val:
                headers['lotte'] = cl
            elif 'Shilla' in val or 'SHILLA' in val:
                headers['shilla'] = cl
            elif 'SSG' in val:
                headers['ssg'] = cl
            elif 'Hyundai' in val or 'HYUNDAI' in val:
                headers['hyundai'] = cl
            elif 'JDC' in val:
                headers['jdc'] = cl

    ean_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ean_idx = ord(headers['ean']) - ord('A')
        ean_val = row[ean_idx] if ean_idx < len(row) else None
        if ean_val is None:
            continue
        ean_str = str(int(ean_val)) if isinstance(ean_val, (int, float)) else str(ean_val).strip()
        def get_val(cl):
            idx = ord(cl) - ord('A')
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else ''
        ean_map[ean_str] = {
            'lotte': get_val(headers['lotte']),
            'shilla': get_val(headers['shilla']),
            'ssg': get_val(headers['ssg']),
            'hyundai': get_val(headers['hyundai']),
            'jdc': get_val(headers['jdc']),
        }
    return ean_map


def parse_invoice_pdf(pdf_path):
    if fitz is None:
        raise ImportError("PyMuPDF not installed")
    doc = fitz.open(pdf_path)
    all_items = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        words = page.get_text("words")
        rows = {}
        for w in words:
            x0, y0, x1, y1, word, *_ = w
            row_key = round(y0 / 5) * 5
            if row_key not in rows:
                rows[row_key] = []
            rows[row_key].append((x0, word))
        for y in sorted(rows.keys()):
            items = sorted(rows[y], key=lambda t: t[0])
            words_only = [item[1] for item in items]
            first = words_only[0].replace(',', '')
            if re.match(r'^\d{13}$', first):
                if len(words_only) < 9:
                    continue
                suffix = words_only[-8:]
                desc_words = words_only[1:-8]
                description = ' '.join(desc_words)
                try:
                    qty = int(float(suffix[2].replace(',', '')))
                except:
                    qty = 0
                try:
                    unit_price = float(suffix[3].replace(',', ''))
                except:
                    unit_price = 0.0
                try:
                    total_price = float(suffix[4].replace(',', ''))
                except:
                    total_price = 0.0
                all_items.append({
                    'ean': first,
                    'description': description,
                    'quantity': qty,
                    'unit_price': unit_price,
                    'total_price': total_price,
                })
    doc.close()
    return all_items


def create_excel(items, pdf_path, retailer_id, output_dir=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    label = RETAILER_MAP[retailer_id][2]

    # Styles
    hdr_font = openpyxl.styles.Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hdr_fill = openpyxl.styles.PatternFill('solid', fgColor='2F5496')
    hdr_align = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side('thin'), right=openpyxl.styles.Side('thin'),
        top=openpyxl.styles.Side('thin'), bottom=openpyxl.styles.Side('thin'))
    num_align = openpyxl.styles.Alignment(horizontal='right', vertical='center')
    sum_fill = openpyxl.styles.PatternFill('solid', fgColor='D6E4F0')

    cols = [
        ('EAN Number', 24),
        (label, 20),
        ('Description', 54),
        ('Quantity', 12),
        ('Unit Price (USD)', 16),
        ('Total Price (USD)', 18),
    ]
    for ci, (txt, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=txt)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border
        ws.column_dimensions[chr(64+ci)].width = w

    for ri, item in enumerate(items, 2):
        ws.cell(ri, 1, item['ean']).border = thin_border
        ws.cell(ri, 2, item['retailer_code']).border = thin_border
        ws.cell(ri, 3, item['description']).border = thin_border
        for ci, val in [(4, item['quantity']), (5, item['unit_price']), (6, item['total_price'])]:
            c = ws.cell(ri, ci, val)
            c.border = thin_border
            c.alignment = num_align

    sr = len(items) + 2
    ws.cell(sr, 1, 'TOTAL').border = thin_border
    ws.cell(sr, 1).fill = sum_fill
    for ci in range(2, 7):
        ws.cell(sr, ci).fill = sum_fill
        ws.cell(sr, ci).border = thin_border
    for ci in [4, 5, 6]:
        c = ws.cell(sr, ci)
        c.value = f'=SUM({chr(64+ci)}2:{chr(64+ci)}{len(items)+1})'
        c.alignment = num_align

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(items)+1}'

    pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
    out_dir = output_dir or resolve_output_dir()
    out = str(Path(out_dir) / f'{pdf_name}_processed.xlsx')
    wb.save(out)
    return out


def register_sku_master(file_path):
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    SKU_STORAGE.mkdir(parents=True, exist_ok=True)

    src_name = os.path.basename(file_path)
    if not src_name.startswith('SKU master file'):
        from datetime import datetime
        ts = datetime.now().strftime('%Y.%m')
        dest_name = f'SKU master file_{ts}.xlsx'
    else:
        dest_name = src_name

    dest = SKU_STORAGE / dest_name

    # Clean old
    for f in SKU_STORAGE.iterdir():
        if f.name.startswith('SKU master file') and f.suffix == '.xlsx' and f.name != dest_name:
            f.unlink(missing_ok=True)

    import shutil
    shutil.copy2(file_path, str(dest))

    cfg = {'sku_master_path': str(dest)}
    with open(CONFIG_FILE, 'w') as f:
        json.dump(cfg, f, indent=2)

    return dest_name


# ─── GUI ───────────────────────────────────────────────────

class InvoiceProcessorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Invoice Processor")
        self.root.geometry("560x520")
        self.root.minsize(480, 420)
        self.root.configure(bg='#f0f2f5')

        # Windows: register for file drag-and-drop (shell32 DragAcceptFiles)
        if sys.platform == 'win32':
            self._enable_drag_drop()

        # Try to set icon
        try:
            self.root.iconbitmap(default='appicon.ico')
        except:
            pass

        self.sku_name = tk.StringVar(value=self._load_sku_name())
        self.status_text = tk.StringVar()
        self.register_info = tk.StringVar()

        self._build_ui()
        self._center_window()

    def _load_sku_name(self):
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE) as f:
                    cfg = json.load(f)
                p = cfg.get('sku_master_path', '')
                if p and os.path.exists(p):
                    return os.path.basename(p)
            except:
                pass
        path = resolve_sku_master()
        if path:
            return os.path.basename(path)
        return ""

    def _load_output_dir(self):
        return resolve_output_dir()

    def _enable_drag_drop(self):
        """Windows: WM_DROPFILES로 파일 드래그 앤 드롭 활성화"""
        try:
            import ctypes
            from ctypes import wintypes

            hwnd = ctypes.windll.user32.GetParent(self.root.winfo_id())
            ctypes.windll.shell32.DragAcceptFiles(hwnd, True)

            # Subclass the window to intercept WM_DROPFILES
            WM_DROPFILES = 0x0233
            GWL_WNDPROC = -4

            # Keep reference to prevent garbage collection
            self._old_wndproc = ctypes.windll.user32.GetWindowLongPtrW(hwnd, GWL_WNDPROC)

            drop_handler = ctypes.CFUNCTYPE(wintypes.LPARAM, wintypes.HWND, wintypes.UINT,
                                            wintypes.WPARAM, wintypes.LPARAM)

            def wndproc(hWnd, msg, wParam, lParam):
                if msg == WM_DROPFILES:
                    hDrop = wParam
                    buffer = ctypes.create_unicode_buffer(1024)
                    count = ctypes.windll.shell32.DragQueryFileW(hDrop, 0xFFFFFFFF, None, 0)
                    paths = []
                    for i in range(count):
                        ctypes.windll.shell32.DragQueryFileW(hDrop, i, buffer, len(buffer))
                        paths.append(buffer.value)
                    ctypes.windll.shell32.DragFinish(hDrop)
                    if paths:
                        self.root.after(0, lambda p=paths: self._handle_dropped_files(p))
                    return 0
                return ctypes.windll.user32.CallWindowProcW(
                    self._old_wndproc, hWnd, msg, wParam, lParam)

            self._drop_proc = drop_handler(wndproc)
            ctypes.windll.user32.SetWindowLongPtrW(hwnd, GWL_WNDPROC,
                                                   ctypes.cast(self._drop_proc, ctypes.c_longlong))
        except Exception as e:
            print(f"⚠️ Drag-drop init failed: {e}")

    def _handle_dropped_files(self, paths):
        pdfs = [p for p in paths if p.lower().endswith('.pdf')]
        xlsx = [p for p in paths if p.lower().endswith('.xlsx')]
        if xlsx:
            self._register_sku(xlsx[0])
        elif len(pdfs) == 1:
            self._handle_file(pdfs[0])
        elif len(pdfs) > 1:
            self._process_batch(pdfs)

    def _build_ui(self):
        # ── Header ──
        header = tk.Frame(self.root, bg='#ffffff', height=48)
        header.pack(fill='x')
        header.pack_propagate(False)

        tk.Label(header, text="📄  Invoice Processor", font=('Segoe UI', 13, 'bold'),
                 bg='#ffffff', fg='#1a1a2e').pack(side='left', padx=16, pady=10)

        # ── Main area ──
        main = tk.Frame(self.root, bg='#f0f2f5')
        main.pack(fill='both', expand=True, padx=24, pady=20)

        # Drop zone
        self.drop_frame = tk.Frame(main, bg='#ffffff', highlightbackground='#c0c4cc',
                                   highlightthickness=2, cursor='hand2')
        self.drop_frame.pack(fill='both', expand=True)
        self.drop_frame.pack_propagate(False)

        self.drop_inner = tk.Frame(self.drop_frame, bg='#ffffff')
        self.drop_inner.place(relx=0.5, rely=0.45, anchor='center')

        self.drop_icon = tk.Label(self.drop_inner, text="📥", font=('Segoe UI', 36),
                                  bg='#ffffff', fg='#888')
        self.drop_icon.pack()

        tk.Label(self.drop_inner, text="Drop invoice PDF here",
                 font=('Segoe UI', 13), bg='#ffffff', fg='#333').pack(pady=(4, 0))
        tk.Label(self.drop_inner, text="or drop SKU master Excel (.xlsx) to register",
                 font=('Segoe UI', 9), bg='#ffffff', fg='#999').pack()
        tk.Label(self.drop_inner, text="or click to select a file",
                 font=('Segoe UI', 8), bg='#ffffff', fg='#bbb').pack(pady=(4, 0))

        # Hover effects
        self.drop_frame.bind('<Enter>', lambda e: self._on_drop_hover(True))
        self.drop_frame.bind('<Leave>', lambda e: self._on_drop_hover(False))
        self.drop_frame.bind('<Button-1>', lambda e: self._select_file())

        # Progress area
        self.progress_frame = tk.Frame(main, bg='#f0f2f5')
        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='indeterminate', length=300)
        self.progress_label = tk.Label(self.progress_frame, text="", font=('Segoe UI', 10),
                                       bg='#f0f2f5', fg='#555')

        # Result area
        self.result_frame = tk.Frame(main, bg='#f0f2f5')

        # ── SKU master bar ──
        sku_bar = tk.Frame(self.root, bg='#ffffff', height=36)
        sku_bar.pack(fill='x', side='bottom')
        sku_bar.pack_propagate(False)

        tk.Label(sku_bar, text="📊", font=('Segoe UI', 10), bg='#ffffff').pack(side='left', padx=(12, 2))
        self.sku_label = tk.Label(sku_bar, textvariable=self.sku_name, font=('Segoe UI', 9),
                                  bg='#ffffff', fg='#555', anchor='w')
        self.sku_label.pack(side='left', fill='x', expand=True, padx=(0, 8))

        tk.Button(sku_bar, text="Change...", font=('Segoe UI', 8), bd=0,
                  bg='#e8ecf0', activebackground='#d0d4d8', cursor='hand2',
                  command=self._select_sku_master).pack(side='right', padx=(0, 12))

        # ── Output dir bar ──
        out_bar = tk.Frame(self.root, bg='#f5f6f8', height=26)
        out_bar.pack(fill='x', side='bottom')
        out_bar.pack_propagate(False)

        self.output_dir_var = tk.StringVar(value=self._load_output_dir())
        tk.Label(out_bar, text="💾", font=('Segoe UI', 9), bg='#f5f6f8').pack(side='left', padx=(12, 2))
        self.out_label = tk.Label(out_bar, textvariable=self.output_dir_var, font=('Segoe UI', 8),
                                  bg='#f5f6f8', fg='#888', anchor='w')
        self.out_label.pack(side='left', fill='x', expand=True, padx=(0, 8))

        tk.Button(out_bar, text="Save to...", font=('Segoe UI', 8), bd=0,
                  bg='#e0e4e8', activebackground='#d0d4d8', cursor='hand2',
                  command=self._select_output_dir).pack(side='right', padx=(0, 12))

        # Status bar
        status_bar = tk.Frame(self.root, bg='#e8ecf0', height=28)
        status_bar.pack(fill='x', side='bottom')
        status_bar.pack_propagate(False)
        tk.Label(status_bar, textvariable=self.status_text, font=('Segoe UI', 8),
                 bg='#e8ecf0', fg='#888', anchor='w').pack(side='left', padx=12)

        # Bind keyboard shortcuts
        self.root.bind('<Control-o>', lambda e: self._select_file())
        self.root.bind('<Control-p>', lambda e: self._select_sku_master())

    def _center_window(self):
        self.root.update_idletasks()
        w, h = 560, 520
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.root.geometry(f'{w}x{h}+{x}+{y}')

    def _on_drop_hover(self, enter):
        self.drop_frame.configure(highlightbackground='#4a7cf7' if enter else '#c0c4cc')
        self.drop_icon.configure(fg='#4a7cf7' if enter else '#888')

    def _set_status(self, text):
        self.status_text.set(text)
        self.root.update_idletasks()

    def _select_file(self):
        paths = filedialog.askopenfilenames(
            title="Select invoice PDF(s)",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if paths:
            pdfs = [p for p in paths if p.lower().endswith('.pdf')]
            if len(pdfs) == 1:
                self._handle_file(pdfs[0])
            elif len(pdfs) > 1:
                self._process_batch(pdfs)
            elif paths and not pdfs:
                # Could be xlsx dropped here too (from drag)
                for p in paths:
                    if p.lower().endswith('.xlsx'):
                        self._register_sku(p)
                        break

    def _select_sku_master(self):
        path = filedialog.askopenfilename(
            title="Select SKU master Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self._register_sku(path)

    def _select_output_dir(self):
        path = filedialog.askdirectory(
            title="Select output folder for Excel files",
            initialdir=self._load_output_dir()
        )
        if path:
            set_output_dir(path)
            self.output_dir_var.set(path)

    def _handle_file(self, path):
        ext = os.path.splitext(path)[1].lower()
        if ext == '.pdf':
            self._process_pdf(path)
        elif ext == '.xlsx':
            self._register_sku(path)
        else:
            messagebox.showwarning("Unsupported File", f"Unsupported file type: {ext}\nUse .pdf for invoices or .xlsx for SKU master.")

    def _show_processing(self, show):
        if show:
            self.drop_frame.pack_forget()
            self.result_frame.pack_forget()
            self.progress_frame.pack(fill='both', expand=True, pady=60)
            self.progress_label.pack(pady=(0, 16))
            self.progress_bar.pack()
            self.progress_bar.start(15)
        else:
            self.progress_bar.stop()
            self.progress_frame.pack_forget()

    def _show_result(self, items, matched, retailer_id, output_path):
        self.progress_frame.pack_forget()
        for w in self.result_frame.winfo_children():
            w.destroy()

        label = RETAILER_MAP[retailer_id][2]
        total_qty = sum(i['quantity'] for i in items)
        total_amt = sum(i['total_price'] for i in items)

        self.result_frame.pack(fill='both', expand=True)

        tk.Label(self.result_frame, text="✅", font=('Segoe UI', 36), bg='#f0f2f5').pack(pady=(16, 4))
        tk.Label(self.result_frame, text="Conversion Complete",
                 font=('Segoe UI', 16, 'bold'), bg='#f0f2f5', fg='#1a1a2e').pack()

        info_frame = tk.Frame(self.result_frame, bg='#ffffff', highlightbackground='#ddd',
                              highlightthickness=1)
        info_frame.pack(pady=20, padx=32, fill='x')

        rows_data = [
            ("🏪  Retailer", label),
            ("📦  Items", f"{len(items)}"),
            ("🔗  SKU Matched", f"{matched}/{len(items)}"),
            ("📊  Total Quantity", f"{total_qty:,}"),
            ("💰  Total Price USD", f"${total_amt:,.2f}"),
        ]
        for i, (lbl, val) in enumerate(rows_data):
            row = tk.Frame(info_frame, bg='#ffffff')
            row.pack(fill='x', padx=16, pady=4)
            tk.Label(row, text=lbl, font=('Segoe UI', 10), bg='#ffffff', fg='#888',
                     anchor='w').pack(side='left')
            tk.Label(row, text=val, font=('Segoe UI', 10, 'bold'), bg='#ffffff', fg='#333',
                     anchor='e').pack(side='right', fill='x', expand=True)
            if i < len(rows_data) - 1:
                ttk.Separator(info_frame, orient='horizontal').pack(fill='x', padx=12)

        btn_frame = tk.Frame(self.result_frame, bg='#f0f2f5')
        btn_frame.pack(pady=12)

        def open_file():
            if output_path and os.path.exists(output_path):
                if sys.platform == 'win32':
                    os.startfile(output_path)
                else:
                    subprocess.run(['open', output_path])

        def open_folder():
            if output_path:
                folder = os.path.dirname(output_path)
                if sys.platform == 'win32':
                    subprocess.run(['explorer', '/select,', output_path])
                else:
                    subprocess.run(['open', '-R', output_path])

        tk.Button(btn_frame, text="📗  Open Excel", font=('Segoe UI', 10, 'bold'),
                  bg='#2F5496', fg='white', bd=0, padx=20, pady=6, cursor='hand2',
                  activebackground='#1e3a6e', command=open_file).pack(side='left', padx=4)

        tk.Button(btn_frame, text="📂  Show in Folder", font=('Segoe UI', 10),
                  bg='#e8ecf0', bd=0, padx=16, pady=6, cursor='hand2',
                  command=open_folder).pack(side='left', padx=4)

        tk.Button(btn_frame, text="🔄  Process Another", font=('Segoe UI', 10),
                  bg='#e8ecf0', bd=0, padx=16, pady=6, cursor='hand2',
                  command=self._reset).pack(side='left', padx=4)

        self._set_status(f"✅ {len(items)} items · {label} · ${total_amt:,.2f}")

    def _show_sku_registered(self, name):
        self.progress_frame.pack_forget()
        for w in self.result_frame.winfo_children():
            w.destroy()
        self.result_frame.pack(fill='both', expand=True)

        tk.Label(self.result_frame, text="📋", font=('Segoe UI', 36), bg='#f0f2f5').pack(pady=(16, 4))
        tk.Label(self.result_frame, text="SKU Master Registered",
                 font=('Segoe UI', 16, 'bold'), bg='#f0f2f5', fg='#1a1a2e').pack()

        tk.Label(self.result_frame, text=f"\"{name}\"",
                 font=('Segoe UI', 11), bg='#f0f2f5', fg='#555').pack(pady=8)

        tk.Label(self.result_frame, text="Ready to process invoices",
                 font=('Segoe UI', 10), bg='#f0f2f5', fg='#888').pack()

        tk.Button(self.result_frame, text="📄  Process a PDF",
                  font=('Segoe UI', 10, 'bold'), bg='#2F5496', fg='white', bd=0,
                  padx=24, pady=8, cursor='hand2', activebackground='#1e3a6e',
                  command=self._select_file).pack(pady=20)

    def _reset(self):
        self.result_frame.pack_forget()
        self.progress_frame.pack_forget()
        self.drop_frame.pack(fill='both', expand=True)
        self._set_status("")

    def _process_pdf(self, path):
        self._show_processing(True)
        self._set_status(f"Processing {os.path.basename(path)}...")

        def work():
            try:
                sku_path = resolve_sku_master()
                if not sku_path:
                    self.root.after(0, lambda: self._show_error("SKU master not found.\nDrag an Excel file to register one first."))
                    return
                retailer_id = detect_retailer(path)
                ean_map = load_sku_master(sku_path)
                items = parse_invoice_pdf(path)
                if not items:
                    self.root.after(0, lambda: self._show_error("No data could be extracted from this PDF."))
                    return
                matched = 0
                for item in items:
                    ean = item['ean']
                    if ean in ean_map:
                        item['retailer_code'] = ean_map[ean][retailer_id]
                        matched += 1
                    else:
                        item['retailer_code'] = ''
                output = create_excel(items, path, retailer_id, resolve_output_dir())
                self.root.after(0, lambda: self._show_result(items, matched, retailer_id, output))
            except Exception as e:
                self.root.after(0, lambda: self._show_error(str(e)))

        threading.Thread(target=work, daemon=True).start()

    def _process_batch(self, paths):
        self._show_processing(True)
        self._set_status(f"Processing 1/{len(paths)}...")

        def work():
            try:
                sku_path = resolve_sku_master()
                if not sku_path:
                    self.root.after(0, lambda: self._show_error("SKU master not found."))
                    return
                ean_map = load_sku_master(sku_path)
                total_items = 0
                total_matched = 0
                total_qty = 0
                total_price = 0.0
                last_output = None
                file_count = 0
                file_results = []

                for i, path in enumerate(paths):
                    self.root.after(0, lambda i=i: self._set_status(f"Processing {i+1}/{len(paths)}..."))
                    retailer_id = detect_retailer(path)
                    items = parse_invoice_pdf(path)
                    if not items:
                        continue
                    matched = 0
                    for item in items:
                        ean = item['ean']
                        if ean in ean_map:
                            item['retailer_code'] = ean_map[ean][retailer_id]
                            matched += 1
                        else:
                            item['retailer_code'] = ''
                    output = create_excel(items, path, retailer_id, resolve_output_dir())
                    total_items += len(items)
                    total_matched += matched
                    total_qty += sum(it['quantity'] for it in items)
                    total_price += sum(it['total_price'] for it in items)
                    last_output = output
                    file_count += 1
                    # Store individual file result
                    file_results.append({
                        'file': os.path.basename(path),
                        'items': len(items),
                        'matched': matched,
                        'qty': sum(it['quantity'] for it in items),
                        'amt': sum(it['total_price'] for it in items),
                        'path': output,
                    })

                self.root.after(0, lambda fr=file_results: self._show_batch_result(
                    total_items, total_matched, total_qty, total_price, file_count, last_output, fr))
            except Exception as e:
                self.root.after(0, lambda: self._show_error(str(e)))

        threading.Thread(target=work, daemon=True).start()

    def _show_batch_result(self, items, matched, qty, amt, files, output_path, file_results=None):
        self.progress_frame.pack_forget()
        for w in self.result_frame.winfo_children():
            w.destroy()
        self.result_frame.pack(fill='both', expand=True)

        tk.Label(self.result_frame, text="✅", font=('Segoe UI', 36), bg='#f0f2f5').pack(pady=(12, 2))

        # Use Notebook for tabs
        if file_results:
            notebook = ttk.Notebook(self.result_frame)
            notebook.pack(pady=(6, 16), padx=32, fill='both', expand=True)

            for fr in file_results:
                file_frame = tk.Frame(notebook, bg='#f0f2f5')
                tab_name = os.path.splitext(fr['file'])[0]
                if tab_name.endswith(' CI'):
                    tab_name = tab_name[:-3]
                notebook.add(file_frame, text=tab_name)

                card = tk.Frame(file_frame, bg='#ffffff', highlightbackground='#ddd', highlightthickness=1)
                card.pack(pady=16, padx=24, fill='x')

                file_rows = [
                    ("📄  File", fr['file']),
                    ("📦  Items", f"{fr['items']}"),
                    ("🔗  SKU Matched", f"{fr['matched']}/{fr['items']}"),
                    ("📊  Quantity", f"{fr['qty']:,}"),
                    ("💰  Total Price USD", f"${fr['amt']:,.2f}"),
                ]
                for lbl, val in file_rows:
                    row = tk.Frame(card, bg='#ffffff')
                    row.pack(fill='x', padx=16, pady=3)
                    tk.Label(row, text=lbl, font=('Segoe UI', 10), bg='#ffffff', fg='#888',
                             anchor='w').pack(side='left')
                    tk.Label(row, text=val, font=('Segoe UI', 10, 'bold'), bg='#ffffff', fg='#333',
                             anchor='e').pack(side='right')

                def make_opener(p=fr['path']):
                    def open_it():
                        if sys.platform == 'win32':
                            os.startfile(p)
                        else:
                            subprocess.run(['open', p])
                    return open_it

                tk.Button(card, text="📗  Open Excel", font=('Segoe UI', 10, 'bold'),
                          bg='#2F5496', fg='white', bd=0, padx=16, pady=5, cursor='hand2',
                          activebackground='#1e3a6e', command=make_opener()).pack(pady=(8, 4))

        btn_frame = tk.Frame(self.result_frame, bg='#f0f2f5')
        btn_frame.pack(pady=12)

        def open_folder():
            if output_path:
                folder = os.path.dirname(output_path)
                if sys.platform == 'win32':
                    subprocess.run(['explorer', folder])
                else:
                    subprocess.run(['open', folder])

        tk.Button(btn_frame, text="📂  Show in Finder", font=('Segoe UI', 10),
                  bg='#2F5496', fg='white', bd=0, padx=20, pady=6, cursor='hand2',
                  activebackground='#1e3a6e', command=open_folder).pack(side='left', padx=4)
        tk.Button(btn_frame, text="🔄  Process More", font=('Segoe UI', 10),
                  bg='#e8ecf0', bd=0, padx=16, pady=6, cursor='hand2',
                  command=self._reset).pack(side='left', padx=4)

        self._set_status(f"✅ {files} files · {items} items · ${amt:,.2f}")

    def _register_sku(self, path):
        self._show_processing(True)
        self._set_status("Registering SKU master...")

        def work():
            try:
                name = register_sku_master(path)
                self.sku_name.set(name)
                self.root.after(0, lambda: self._show_sku_registered(name))
            except Exception as e:
                self.root.after(0, lambda: self._show_error(str(e)))

        threading.Thread(target=work, daemon=True).start()

    def _show_error(self, msg):
        self._show_processing(False)
        self.drop_frame.pack(fill='both', expand=True)
        messagebox.showerror("Error", msg)
        self._set_status("❌ " + msg.split('\n')[0])

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    # Check dependencies
    missing = []
    if fitz is None:
        missing.append("PyMuPDF (pip install PyMuPDF)")
    if openpyxl is None:
        missing.append("openpyxl (pip install openpyxl)")
    if missing:
        print("Missing dependencies:")
        for m in missing:
            print(f"  {m}")
        input("\nPress Enter to exit...")
        sys.exit(1)

    # Check for command-line args (drag onto .exe in Windows Explorer)
    if len(sys.argv) > 1:
        # Process files directly, show messagebox at end
        results = []
        pdfs = [f for f in sys.argv[1:] if f.lower().endswith('.pdf')]
        xlsx = [f for f in sys.argv[1:] if f.lower().endswith('.xlsx')]
        if xlsx:
            name = register_sku_master(xlsx[0])
            results.append(f"SKU master registered: {name}")
        if pdfs:
            sku_path = resolve_sku_master()
            if not sku_path:
                results.append("SKU master not found.")
            else:
                ean_map = load_sku_master(sku_path)
                for pdf in pdfs:
                    rid = detect_retailer(pdf)
                    items = parse_invoice_pdf(pdf)
                    matched = 0
                    for item in items:
                        ean = item['ean']
                        if ean in ean_map:
                            item['retailer_code'] = ean_map[ean][rid]
                            matched += 1
                        else:
                            item['retailer_code'] = ''
                    out = create_excel(items, pdf, rid, resolve_output_dir())
                    results.append(f"{os.path.basename(pdf)}\n  {len(items)} items, {matched} matched\n  → {os.path.basename(out)}")
        if not pdfs and not xlsx:
            results.append("Unsupported file.\nUse .pdf (invoice) or .xlsx (SKU master)")

        # Show result via tkinter messagebox (works with --windowed)
        try:
            import tkinter.messagebox as mb
            root = tk.Tk()
            root.withdraw()
            mb.showinfo("Invoice Processor", "\n\n".join(results))
            root.destroy()
        except:
            print("\n\n".join(results))
            input("\nPress Enter...")
        sys.exit(0)
    else:
        app = InvoiceProcessorApp()
        app.run()
